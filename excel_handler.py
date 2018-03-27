import os

import yaml
import xlrd

from openpyxl import load_workbook
from util_func import securely_check_dir


class ExcelHandler:
    def __init__(self, config):
        self.config = config

        securely_check_dir('forms')
        securely_check_dir('att')
        securely_check_dir('config')

        self.subject = []
        for item in self.config['responds'].keys():
            if not item.startswith('default'):
                self.subject.append(item)

        self.handle_config = []
        config_root = 'config'
        for _, _, files in os.walk(config_root):
            for file in files:
                subject, _ = os.path.splitext(file)
                if subject != 'top' and not subject.endswith('-old'):
                    with open(os.path.join(config_root, file)) as fp:
                        subject_config = yaml.load(fp.read())
                    self.handle_config.append({'subject_name': subject, 'config': subject_config})

    def handle(self):
        att_root = 'att'
        for subject_config in self.handle_config:
            subject = subject_config['subject_name']
            config = subject_config['config']
            if os.path.exists(os.path.join(att_root, subject)):
                for _, _, files in os.walk(os.path.join(att_root, subject)):
                    for f in files:
                        short_name, ext = os.path.splitext(f)
                        if not short_name.endswith('-old') and not f.startswith('.'):
                            workbook = load_workbook(os.path.join(att_root, subject, f))
                            sheet_names = config.keys()
                            for sheet_name in sheet_names:
                                from_row = config[sheet_name]['header']['row']['to'] + 1
                                from_column = config[sheet_name]['column']['from']
                                sheet = workbook[sheet_name]
                                content = []
                                tmp_work_book = xlrd.open_workbook(os.path.join(att_root, subject, f))
                                tmp_sheet = tmp_work_book.sheet_by_name(sheet_name)
                                lines = tmp_sheet.nrows
                                tmp_work_book.release_resources()
                                for i in range(from_row, lines + 1):
                                    row = [val.value for val in sheet[i]][from_column - 1:-1]
                                    content.append(row)
                                form_workbook = load_workbook(
                                    os.path.join('forms', subject, config[sheet_name]['destination_file']))
                                form_sheet = form_workbook[sheet_name]
                                tmp_work_book = xlrd.open_workbook(
                                    os.path.join('forms', subject, config[sheet_name]['destination_file']))
                                tmp_sheet = tmp_work_book.sheet_by_name(sheet_name)
                                lines = tmp_sheet.nrows
                                tmp_work_book.release_resources()
                                for i in range(len(content)):
                                    for j in range(len(content[i])):
                                        form_sheet[lines + i + 1][j].value = content[i][j]
                                form_workbook.save(
                                    os.path.join('forms', subject, config[sheet_name]['destination_file']))
                                form_workbook.close()
                                workbook.close()
                            os.rename(os.path.join(att_root, subject, f),
                                      os.path.join(att_root, subject, '{}{}{}'.format(short_name, '-old', ext)))


if __name__ == '__main__':
    config_file = 'config/top.yml'
    if not os.path.exists(config_file):
        print('No top.yml file found!')
        exit(-1)
    with open(config_file, encoding='utf-8') as f:
        config_file = yaml.load(f.read())
    excel_handler = ExcelHandler(config_file)
    excel_handler.handle()
